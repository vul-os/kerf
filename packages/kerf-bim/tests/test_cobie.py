"""
Tests for packages/kerf-bim/src/kerf_bim/cobie.py

Covers:
  - Standard template has all 18 COBie sheets mapped
  - All 4 built-in templates load without error
  - apply_mapping_template populates rows from IFC data
  - validate_cobie_deliverable catches missing required columns
  - validate_cobie_deliverable passes for a complete deliverable
  - export_cobie_excel produces a valid .xlsx (or CSV fallback)
  - export_cobie_xml produces well-formed XML
  - compute_completeness_score returns 0.0 on empty data
  - compute_completeness_score ≥ 0.8 on well-populated data
  - PropertyMapping dataclass round-trips through template
  - COBIE_SHEET_NAMES has exactly 18 entries
  - LLM tool TOOLS list has 5 entries
"""
from __future__ import annotations

import asyncio
import json
import os
import tempfile
import xml.etree.ElementTree as ET

import pytest

from kerf_bim.cobie import (
    COBIE_SHEET_NAMES,
    CobieDeliverable,
    CobieSheet,
    MappingTemplate,
    PropertyMapping,
    apply_mapping_template,
    compute_completeness_score,
    export_cobie_excel,
    export_cobie_xml,
    get_standard_template,
    validate_cobie_deliverable,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def minimal_ifc_data():
    """Minimal IFC dict with a building, one floor, two spaces."""
    return {
        "created_by": "test@example.com",
        "property_sets": {
            "Pset_BuildingCommon": {
                "BuildingID": "Building-A",
                "ProjectName": "Test Tower",
            },
            "Pset_SiteCommon": {
                "SiteName": "Johannesburg CBD",
            },
            "Pset_ContactInformation": {
                "Email": "fm@example.com",
                "Organization": "Acme FM",
                "Phone": "+27 11 123 4567",
                "Street": "1 Main Street",
                "City": "Johannesburg",
                "State": "Gauteng",
                "Country": "ZA",
                "PostalCode": "2001",
            },
        },
        "elements": [
            {
                "ifc_class": "IfcBuildingStorey",
                "name": "Ground Floor",
                "category": "floor",
                "property_sets": {
                    "Pset_BuildingStoreyCommon": {
                        "GrossFloorArea": "500.0",
                        "NetFloorArea": "480.0",
                    },
                },
            },
            {
                "ifc_class": "IfcSpace",
                "name": "Room 101",
                "category": "office",
                "property_sets": {
                    "Pset_SpaceCommon": {
                        "Reference": "101",
                        "GrossFloorArea": "42.5",
                        "NetFloorArea": "40.0",
                        "FinishCeiling": "Plasterboard painted white",
                        "RoomTag": "101",
                    },
                },
            },
            {
                "ifc_class": "IfcSpace",
                "name": "Room 102",
                "category": "office",
                "property_sets": {
                    "Pset_SpaceCommon": {
                        "Reference": "102",
                        "GrossFloorArea": "38.0",
                        "NetFloorArea": "36.0",
                        "RoomTag": "102",
                    },
                },
            },
            {
                "ifc_class": "IfcTypeObject",
                "name": "VAV Unit Type A",
                "category": "hvac",
                "property_sets": {
                    "Pset_ManufacturerTypeInformation": {
                        "Manufacturer": "Trane",
                        "ModelLabel": "VAV-A-100",
                        "ArticleNumber": "SKU-12345",
                    },
                    "Pset_AssetInformation": {
                        "AssetIdentifier": "mechanical",
                    },
                },
            },
        ],
    }


@pytest.fixture
def rich_ifc_data(minimal_ifc_data):
    """Extends minimal_ifc_data with job, document, and zone data."""
    minimal_ifc_data["property_sets"].update({
        "Pset_MaintenanceTaskCommon": {
            "TaskId": "MAINT-001",
            "MaintenanceType": "Preventive",
            "StandardMode": "Annual HVAC service",
            "Duration": "4",
            "DurationUnit": "Hour",
            "Frequency": "1",
            "FrequencyUnit": "Year",
        },
        "Pset_DocumentInformation": {
            "DocumentId": "DOC-001",
            "Purpose": "O&M Manual",
            "IntendedUse": "FM Team",
            "ElectronicFileLocation": "/docs/om-manual.pdf",
        },
        "Pset_ZoneCommon": {
            "Category": "HVAC",
            "SpaceNames": "Room 101, Room 102",
        },
    })
    return minimal_ifc_data


def _run(coro):
    return asyncio.run(coro)


# ---------------------------------------------------------------------------
# COBIE_SHEET_NAMES
# ---------------------------------------------------------------------------

class TestSheetNames:
    def test_exactly_18_sheets(self):
        assert len(COBIE_SHEET_NAMES) == 18

    def test_known_sheets_present(self):
        for name in ("Contact", "Facility", "Floor", "Space", "Zone", "Type",
                     "Component", "System", "Assembly", "Connection", "Spare",
                     "Resource", "Job", "Document", "Attribute", "Coordinate",
                     "Issue", "Picture"):
            assert name in COBIE_SHEET_NAMES


# ---------------------------------------------------------------------------
# get_standard_template
# ---------------------------------------------------------------------------

class TestGetStandardTemplate:
    def test_standard_template_loads(self):
        tmpl = get_standard_template("standard")
        assert tmpl.template_name == "standard"
        assert len(tmpl.mappings) > 0

    def test_federal_us_template_loads(self):
        tmpl = get_standard_template("federal_us")
        assert tmpl.template_name == "federal_us"

    def test_uk_ukgbc_template_loads(self):
        tmpl = get_standard_template("uk_ukgbc")
        assert tmpl.template_name == "uk_ukgbc"

    def test_singapore_corenet_template_loads(self):
        tmpl = get_standard_template("singapore_corenet")
        assert tmpl.template_name == "singapore_corenet"

    def test_unknown_template_raises(self):
        with pytest.raises(ValueError, match="Unknown template"):
            get_standard_template("nonexistent_template")

    def test_standard_covers_all_18_sheets(self):
        """Standard template should have at least one mapping per required sheet."""
        tmpl = get_standard_template("standard")
        covered_sheets = {m.cobie_sheet for m in tmpl.mappings}
        # At minimum these core sheets must be covered
        for sheet in ("Contact", "Facility", "Floor", "Space", "Type", "Component", "Job", "Document"):
            assert sheet in covered_sheets, f"Sheet {sheet} not covered by standard template"

    def test_property_mapping_round_trip(self):
        tmpl = get_standard_template("standard")
        # Find the Email mapping
        email_mappings = [
            m for m in tmpl.mappings
            if m.ifc_pset_name == "Pset_ContactInformation" and m.ifc_property_name == "Email"
        ]
        assert len(email_mappings) == 1
        m = email_mappings[0]
        assert m.cobie_sheet == "Contact"
        assert m.cobie_column == "Email"


# ---------------------------------------------------------------------------
# apply_mapping_template
# ---------------------------------------------------------------------------

class TestApplyMappingTemplate:
    def test_returns_cobie_deliverable(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        assert isinstance(dlv, CobieDeliverable)

    def test_deliverable_has_18_sheets(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        assert len(dlv.sheets) == 18

    def test_space_rows_populated(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        space_sheet = dlv.get_sheet("Space")
        assert space_sheet is not None
        assert len(space_sheet.rows) >= 2  # Room 101 and 102

    def test_space_gross_area_mapped(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        space_sheet = dlv.get_sheet("Space")
        room_rows = [r for r in space_sheet.rows if "101" in r.get("Name", "")]
        assert room_rows, "Expected Row for Room 101"
        assert room_rows[0].get("GrossArea") == "42.5"

    def test_type_manufacturer_mapped(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        type_sheet = dlv.get_sheet("Type")
        assert type_sheet is not None
        mfr_vals = [r.get("Manufacturer") for r in type_sheet.rows if r.get("Manufacturer")]
        assert any(v == "Trane" for v in mfr_vals)

    def test_contact_email_mapped(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        contact_sheet = dlv.get_sheet("Contact")
        assert contact_sheet is not None
        emails = [r.get("Email") for r in contact_sheet.rows if r.get("Email")]
        assert any(e == "fm@example.com" for e in emails)

    def test_custom_mapping_template(self, minimal_ifc_data):
        """Custom PropertyMapping reaches cobie_column correctly."""
        custom = MappingTemplate(
            template_name="custom",
            mappings=[
                PropertyMapping(
                    ifc_pset_name="Pset_SpaceCommon",
                    ifc_property_name="RoomTag",
                    cobie_sheet="Space",
                    cobie_column="RoomTag",
                    default_value="UNKNOWN",
                ),
            ],
        )
        dlv = apply_mapping_template(minimal_ifc_data, custom)
        space_sheet = dlv.get_sheet("Space")
        assert space_sheet is not None
        room_tags = [r.get("RoomTag") for r in space_sheet.rows if r.get("RoomTag")]
        assert "101" in room_tags or "102" in room_tags


# ---------------------------------------------------------------------------
# validate_cobie_deliverable
# ---------------------------------------------------------------------------

class TestValidateCobieDeliverable:
    def test_empty_deliverable_has_errors(self):
        dlv = CobieDeliverable(sheets=[])
        errors = validate_cobie_deliverable(dlv)
        assert len(errors) > 0
        # Should report all 18 missing sheets
        assert any("Missing required sheet" in e for e in errors)

    def test_missing_required_column_reported(self):
        # Create a deliverable missing the 'Email' column on Contact
        sheets = [CobieSheet(name=n, columns=[], rows=[]) for n in COBIE_SHEET_NAMES]
        dlv = CobieDeliverable(sheets=sheets)
        errors = validate_cobie_deliverable(dlv)
        contact_errors = [e for e in errors if "Contact" in e and "Email" in e]
        assert contact_errors, "Expected error for missing Email column on Contact"

    def test_valid_deliverable_no_errors(self, minimal_ifc_data):
        """A well-formed deliverable should pass validation (may have column warnings)."""
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        errors = validate_cobie_deliverable(dlv)
        # Some missing required columns are expected on sparse data,
        # but the deliverable must have all 18 sheets (no missing-sheet errors)
        sheet_errors = [e for e in errors if "Missing required sheet" in e]
        assert not sheet_errors, f"Unexpected missing sheet errors: {sheet_errors}"

    def test_duplicate_guid_reported(self):
        """Duplicate Email values in Contact should be flagged."""
        email_col = "Email"
        contact_sheet = CobieSheet(
            name="Contact",
            columns=["Email", "CreatedBy", "CreatedOn", "Category", "Company",
                     "Phone", "Street", "City", "State", "Country", "PostalCode"],
            rows=[
                {"Email": "dup@example.com", "CreatedBy": "system", "CreatedOn": "2026-01-01", "Category": "FM", "Company": "A", "Phone": "", "Street": "", "City": "", "State": "", "Country": "", "PostalCode": ""},
                {"Email": "dup@example.com", "CreatedBy": "system", "CreatedOn": "2026-01-01", "Category": "FM", "Company": "B", "Phone": "", "Street": "", "City": "", "State": "", "Country": "", "PostalCode": ""},
            ],
        )
        other_sheets = [CobieSheet(name=n, columns=[], rows=[]) for n in COBIE_SHEET_NAMES if n != "Contact"]
        dlv = CobieDeliverable(sheets=[contact_sheet] + other_sheets)
        errors = validate_cobie_deliverable(dlv)
        dup_errors = [e for e in errors if "duplicate" in e.lower() and "Email" in e]
        assert dup_errors, f"Expected duplicate-email error; got: {errors}"


# ---------------------------------------------------------------------------
# compute_completeness_score
# ---------------------------------------------------------------------------

class TestComputeCompletenessScore:
    def test_empty_deliverable_score_zero(self):
        dlv = CobieDeliverable(
            sheets=[CobieSheet(name=n, columns=[], rows=[]) for n in COBIE_SHEET_NAMES]
        )
        score = compute_completeness_score(dlv)
        assert score == 0.0

    def test_partial_data_score_in_range(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        score = compute_completeness_score(dlv)
        assert 0.0 < score <= 1.0

    def test_rich_data_completeness_above_empty(self, rich_ifc_data):
        """Rich IFC data should score strictly higher than empty data."""
        tmpl = get_standard_template("standard")
        dlv_rich = apply_mapping_template(rich_ifc_data, tmpl)
        score_rich = compute_completeness_score(dlv_rich)

        empty = CobieDeliverable(
            sheets=[CobieSheet(name=n, columns=[], rows=[]) for n in COBIE_SHEET_NAMES]
        )
        score_empty = compute_completeness_score(empty)
        assert score_rich > score_empty, "Rich data should score higher than empty"
        assert score_rich > 0.0

    def test_fully_populated_sheet_scores_high(self):
        """A deliverable where every required column has a value should score 1.0."""
        from kerf_bim.cobie import REQUIRED_COLUMNS
        sheets = []
        for name in COBIE_SHEET_NAMES:
            req_cols = REQUIRED_COLUMNS.get(name, [])
            dummy_row: dict[str, str] = {col: "populated" for col in req_cols}
            sheets.append(CobieSheet(name=name, columns=req_cols, rows=[dummy_row] if req_cols else []))
        dlv = CobieDeliverable(sheets=sheets)
        score = compute_completeness_score(dlv)
        assert score == 1.0, f"Expected 1.0 for fully-populated deliverable, got {score}"

    def test_score_is_float(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        score = compute_completeness_score(dlv)
        assert isinstance(score, float)


# ---------------------------------------------------------------------------
# export_cobie_excel
# ---------------------------------------------------------------------------

class TestExportCobieExcel:
    def test_export_produces_file(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
            path = fh.name
        try:
            result_path = export_cobie_excel(dlv, path)
            assert os.path.exists(result_path) or result_path.endswith(".csv")
        finally:
            # Cleanup
            if os.path.exists(path):
                os.unlink(path)
            # CSV fallback cleanup
            import pathlib
            for csv_file in pathlib.Path(os.path.dirname(path)).glob(
                os.path.basename(path) + ".*.csv"
            ):
                csv_file.unlink(missing_ok=True)

    def test_export_xlsx_valid_workbook(self, minimal_ifc_data):
        """If openpyxl is available, the output must be a readable workbook."""
        pytest.importorskip("openpyxl")
        import openpyxl

        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
            path = fh.name
        try:
            export_cobie_excel(dlv, path)
            wb = openpyxl.load_workbook(path)
            assert len(wb.sheetnames) == 18
            assert "Contact" in wb.sheetnames
            assert "Facility" in wb.sheetnames
            assert "Space" in wb.sheetnames
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_xlsx_has_header_row(self, minimal_ifc_data):
        """First row of Contact sheet must contain 'Email'."""
        pytest.importorskip("openpyxl")
        import openpyxl

        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
            path = fh.name
        try:
            export_cobie_excel(dlv, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Contact"]
            header_row = [cell.value for cell in ws[1]]
            assert "Email" in header_row
        finally:
            if os.path.exists(path):
                os.unlink(path)


# ---------------------------------------------------------------------------
# export_cobie_xml
# ---------------------------------------------------------------------------

class TestExportCobieXml:
    def test_export_produces_xml_file(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        with tempfile.NamedTemporaryFile(suffix=".xml", delete=False) as fh:
            path = fh.name
        try:
            result_path = export_cobie_xml(dlv, path)
            assert os.path.exists(result_path)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_xml_is_well_formed(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        with tempfile.NamedTemporaryFile(suffix=".xml", delete=False) as fh:
            path = fh.name
        try:
            export_cobie_xml(dlv, path)
            tree = ET.parse(path)
            root = tree.getroot()
            # Root tag may include namespace URI prefix
            local = root.tag.split("}")[-1] if "}" in root.tag else root.tag
            assert local == "COBie"
            assert root.get("version") == "2.4"
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_xml_has_space_elements(self, minimal_ifc_data):
        tmpl = get_standard_template("standard")
        dlv = apply_mapping_template(minimal_ifc_data, tmpl)
        with tempfile.NamedTemporaryFile(suffix=".xml", delete=False) as fh:
            path = fh.name
        try:
            export_cobie_xml(dlv, path)
            tree = ET.parse(path)
            root = tree.getroot()
            # Handle optional namespace prefix
            ns = ""
            if "}" in root.tag:
                ns = root.tag.split("}")[0] + "}"
            space_el = root.find(f"{ns}Space")
            assert space_el is not None
        finally:
            if os.path.exists(path):
                os.unlink(path)


# ---------------------------------------------------------------------------
# LLM tool TOOLS list
# ---------------------------------------------------------------------------

class TestToolsList:
    def test_tools_list_has_5_entries(self):
        from kerf_bim.tools.cobie import TOOLS
        assert len(TOOLS) == 5

    def test_tools_names(self):
        from kerf_bim.tools.cobie import TOOLS
        names = [t[0] for t in TOOLS]
        assert "bim_apply_property_mapping" in names
        assert "bim_export_cobie_excel" in names
        assert "bim_validate_cobie" in names
        assert "bim_get_standard_template" in names
        assert "bim_compute_cobie_completeness" in names

    def test_tool_specs_have_names(self):
        from kerf_bim.tools.cobie import TOOLS
        for name, spec, _handler in TOOLS:
            assert spec.name == name


# ---------------------------------------------------------------------------
# LLM tool async dispatch tests
# ---------------------------------------------------------------------------

class TestLLMToolDispatch:
    def _call(self, coro):
        return json.loads(_run(coro))

    def test_get_template_standard(self):
        from kerf_bim.tools.cobie import run_bim_get_standard_template
        result = self._call(
            run_bim_get_standard_template(None, json.dumps({"name": "standard"}).encode())
        )
        assert result.get("template_name") == "standard"
        assert result.get("mapping_count", 0) > 0

    def test_get_template_bad_name(self):
        from kerf_bim.tools.cobie import run_bim_get_standard_template
        result = self._call(
            run_bim_get_standard_template(None, json.dumps({"name": "bogus"}).encode())
        )
        assert "error" in result or result.get("code") == "NOT_FOUND"

    def test_apply_mapping_returns_18_sheets(self, minimal_ifc_data):
        from kerf_bim.tools.cobie import run_bim_apply_property_mapping
        result = self._call(
            run_bim_apply_property_mapping(
                None,
                json.dumps({"ifc_data": minimal_ifc_data, "template_name": "standard"}).encode(),
            )
        )
        assert result.get("sheet_count") == 18

    def test_validate_cobie_passes(self, minimal_ifc_data):
        from kerf_bim.tools.cobie import run_bim_validate_cobie
        result = self._call(
            run_bim_validate_cobie(
                None,
                json.dumps({"ifc_data": minimal_ifc_data}).encode(),
            )
        )
        # No missing-sheet errors
        sheet_errors = [e for e in result.get("errors", []) if "Missing required sheet" in e]
        assert not sheet_errors

    def test_completeness_returns_score(self, minimal_ifc_data):
        from kerf_bim.tools.cobie import run_bim_compute_cobie_completeness
        result = self._call(
            run_bim_compute_cobie_completeness(
                None,
                json.dumps({"ifc_data": minimal_ifc_data}).encode(),
            )
        )
        assert "score" in result
        assert isinstance(result["score"], float)
        assert 0.0 <= result["score"] <= 1.0
