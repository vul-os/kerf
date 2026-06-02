"""
cobie.py — COBie 2.4 / 3.0 deliverable builder.

COBie (Construction Operations Building information exchange) maps IFC property
sets to a structured spreadsheet deliverable for Facility Management handoff.

References:
  - COBie 2.4 spec: https://www.nibs.org/COBie
  - BuildingSMART IFC-to-COBie mapping
  - UK BIM Framework COBie schema
  - Singapore BCA CorNet COBie variant

Structures
----------
CobieRow           dict[str, str] keyed by column name
CobieSheet         dataclass: name, columns, rows
CobieDeliverable   dataclass: sheets
PropertyMapping    dataclass: ifc_pset_name, ifc_property_name, cobie_sheet, cobie_column, default_value
MappingTemplate    dataclass: template_name, mappings

Functions
---------
get_standard_template(name)              → MappingTemplate
apply_mapping_template(ifc_data, template) → CobieDeliverable
validate_cobie_deliverable(deliverable)  → list[str]
export_cobie_excel(deliverable, path)    → str
export_cobie_xml(deliverable, path)      → str
compute_completeness_score(deliverable)  → float
"""

from __future__ import annotations

import csv
import io
import json
import uuid
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Any

# ---------------------------------------------------------------------------
# COBie sheet definitions
# ---------------------------------------------------------------------------

#: All 18 COBie 2.4 sheets in canonical order.
COBIE_SHEET_NAMES: list[str] = [
    "Contact",
    "Facility",
    "Floor",
    "Space",
    "Zone",
    "Type",
    "Component",
    "System",
    "Assembly",
    "Connection",
    "Spare",
    "Resource",
    "Job",
    "Document",
    "Attribute",
    "Coordinate",
    "Issue",
    "Picture",
]

#: Required columns per sheet (minimal COBie 2.4 compliance).
REQUIRED_COLUMNS: dict[str, list[str]] = {
    "Contact": ["Email", "CreatedBy", "CreatedOn", "Category", "Company", "Phone", "Street", "City", "State", "Country", "PostalCode"],
    "Facility": ["Name", "CreatedBy", "CreatedOn", "Category", "ProjectName", "SiteName", "LinearUnits", "AreaUnits", "VolumeUnits", "CurrencyUnit", "AreaMeasurement"],
    "Floor": ["Name", "CreatedBy", "CreatedOn", "Category", "Elevation", "Height"],
    "Space": ["Name", "CreatedBy", "CreatedOn", "Category", "FloorName", "Description", "GrossArea", "NetArea", "RoomTag"],
    "Zone": ["Name", "CreatedBy", "CreatedOn", "Category", "SpaceNames"],
    "Type": ["Name", "CreatedBy", "CreatedOn", "Category", "Description", "AssetType", "Manufacturer", "ModelNumber"],
    "Component": ["Name", "CreatedBy", "CreatedOn", "TypeName", "Space", "Description", "TagNumber"],
    "System": ["Name", "CreatedBy", "CreatedOn", "Category", "ComponentNames"],
    "Assembly": ["Name", "CreatedBy", "CreatedOn", "SheetName", "ParentName", "ChildNames", "AssemblyType"],
    "Connection": ["Name", "CreatedBy", "CreatedOn", "ConnectionType", "SheetName", "RowName1", "RowName2"],
    "Spare": ["Name", "CreatedBy", "CreatedOn", "Category", "TypeName", "Suppliers"],
    "Resource": ["Name", "CreatedBy", "CreatedOn", "Category", "Description"],
    "Job": ["Name", "CreatedBy", "CreatedOn", "Category", "Status", "TypeName", "Description", "Duration", "DurationUnit", "Frequency", "FrequencyUnit"],
    "Document": ["Name", "CreatedBy", "CreatedOn", "Category", "ApprovalBy", "Stage", "SheetName", "RowName", "Directory", "File"],
    "Attribute": ["Name", "CreatedBy", "CreatedOn", "Category", "SheetName", "RowName", "Value"],
    "Coordinate": ["Name", "CreatedBy", "CreatedOn", "Category", "SheetName", "RowName", "CoordinateXAxis", "CoordinateYAxis", "CoordinateZAxis"],
    "Issue": ["Name", "CreatedBy", "CreatedOn", "Type", "Risk", "Chance", "Impact", "SheetName1", "RowName1"],
    "Picture": ["Name", "CreatedBy", "CreatedOn", "Category", "SheetName", "RowName", "PictureName"],
}

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

CobieRow = dict[str, str]


@dataclass
class CobieSheet:
    """A single COBie spreadsheet tab."""

    name: str
    columns: list[str]
    rows: list[CobieRow] = field(default_factory=list)


@dataclass
class CobieDeliverable:
    """Full COBie workbook deliverable."""

    sheets: list[CobieSheet] = field(default_factory=list)

    def get_sheet(self, name: str) -> CobieSheet | None:
        for s in self.sheets:
            if s.name == name:
                return s
        return None


@dataclass
class PropertyMapping:
    """Maps one IFC property → one COBie cell."""

    ifc_pset_name: str
    ifc_property_name: str
    cobie_sheet: str
    cobie_column: str
    default_value: str = ""


@dataclass
class MappingTemplate:
    """Named collection of IFC→COBie property mappings."""

    template_name: str
    mappings: list[PropertyMapping] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Standard templates
# ---------------------------------------------------------------------------

def _standard_template() -> MappingTemplate:
    """COBie 2.4 standard / generic template."""
    return MappingTemplate(
        template_name="standard",
        mappings=[
            # Contact
            PropertyMapping("Pset_ContactInformation", "Email",   "Contact", "Email"),
            PropertyMapping("Pset_ContactInformation", "Organization", "Contact", "Company"),
            PropertyMapping("Pset_ContactInformation", "Phone",   "Contact", "Phone"),
            PropertyMapping("Pset_ContactInformation", "Street",  "Contact", "Street"),
            PropertyMapping("Pset_ContactInformation", "City",    "Contact", "City"),
            PropertyMapping("Pset_ContactInformation", "State",   "Contact", "State"),
            PropertyMapping("Pset_ContactInformation", "Country", "Contact", "Country"),
            PropertyMapping("Pset_ContactInformation", "PostalCode", "Contact", "PostalCode"),
            # Facility
            PropertyMapping("Pset_BuildingCommon", "BuildingID",    "Facility", "Name"),
            PropertyMapping("Pset_BuildingCommon", "ProjectName",   "Facility", "ProjectName"),
            PropertyMapping("Pset_SiteCommon",     "SiteName",      "Facility", "SiteName"),
            PropertyMapping("Pset_BuildingCommon", "GrossPlannedArea", "Facility", "GrossArea"),
            # Floor (IfcBuildingStorey)
            PropertyMapping("Pset_BuildingStoreyCommon", "EntranceLevel",  "Floor", "Name"),
            PropertyMapping("Pset_BuildingStoreyCommon", "GrossFloorArea", "Floor", "GrossArea"),
            PropertyMapping("Pset_BuildingStoreyCommon", "NetFloorArea",   "Floor", "NetArea"),
            # Space (IfcSpace)
            PropertyMapping("Pset_SpaceCommon", "Reference",       "Space", "Name"),
            PropertyMapping("Pset_SpaceCommon", "IsExternal",      "Space", "Category"),
            PropertyMapping("Pset_SpaceCommon", "GrossFloorArea",  "Space", "GrossArea"),
            PropertyMapping("Pset_SpaceCommon", "NetFloorArea",    "Space", "NetArea"),
            PropertyMapping("Pset_SpaceCommon", "FinishCeiling",   "Space", "Description"),
            PropertyMapping("Pset_SpaceCommon", "RoomTag",         "Space", "RoomTag"),
            # Zone
            PropertyMapping("Pset_ZoneCommon", "Category",  "Zone", "Category"),
            PropertyMapping("Pset_ZoneCommon", "SpaceNames","Zone", "SpaceNames"),
            # Type (IfcTypeObject)
            PropertyMapping("Pset_ManufacturerTypeInformation", "Manufacturer",  "Type", "Manufacturer"),
            PropertyMapping("Pset_ManufacturerTypeInformation", "ModelLabel",    "Type", "ModelNumber"),
            PropertyMapping("Pset_ManufacturerTypeInformation", "ArticleNumber", "Type", "Description"),
            PropertyMapping("Pset_AssetInformation",            "AssetIdentifier","Type", "AssetType"),
            # Component
            PropertyMapping("Pset_ManufacturerTypeInformation", "Manufacturer",   "Component", "Description"),
            PropertyMapping("Pset_ComponentInformation",        "TagNumber",       "Component", "TagNumber"),
            # System
            PropertyMapping("Pset_SystemCommon", "Category",       "System", "Category"),
            # Job / maintenance
            PropertyMapping("Pset_MaintenanceTaskCommon", "TaskId",          "Job", "Name"),
            PropertyMapping("Pset_MaintenanceTaskCommon", "MaintenanceType", "Job", "Category"),
            PropertyMapping("Pset_MaintenanceTaskCommon", "StandardMode",    "Job", "Description"),
            PropertyMapping("Pset_MaintenanceTaskCommon", "Duration",        "Job", "Duration"),
            PropertyMapping("Pset_MaintenanceTaskCommon", "DurationUnit",    "Job", "DurationUnit"),
            # Document
            PropertyMapping("Pset_DocumentInformation", "DocumentId",    "Document", "Name"),
            PropertyMapping("Pset_DocumentInformation", "Purpose",       "Document", "Category"),
            PropertyMapping("Pset_DocumentInformation", "IntendedUse",   "Document", "ApprovalBy"),
            PropertyMapping("Pset_DocumentInformation", "ElectronicFileLocation", "Document", "File"),
        ],
    )


def _federal_us_template() -> MappingTemplate:
    """US Federal (GSA/USACE) COBie template — extends standard with federal psets."""
    base = _standard_template()
    extra = [
        PropertyMapping("Pset_FederalAgency",      "AgencyName",    "Facility", "Description"),
        PropertyMapping("Pset_FederalAgency",      "ProjectNumber", "Facility", "ProjectName"),
        PropertyMapping("Pset_FederalAgency",      "ContractNumber","Facility", "SiteName"),
        PropertyMapping("Pset_COBie_Component",    "TagNumber",     "Component", "TagNumber"),
        PropertyMapping("Pset_COBie_Component",    "AssetIdentifier","Component", "Name"),
        PropertyMapping("Pset_MaintenanceTaskCommon", "Frequency",     "Job", "Frequency"),
        PropertyMapping("Pset_MaintenanceTaskCommon", "FrequencyUnit", "Job", "FrequencyUnit"),
        PropertyMapping("Pset_WarrantyCommon",     "WarrantyPeriod", "Spare", "Name"),
        PropertyMapping("Pset_WarrantyCommon",     "WarrantyDescription", "Spare", "Category"),
    ]
    return MappingTemplate(
        template_name="federal_us",
        mappings=base.mappings + extra,
    )


def _uk_ukgbc_template() -> MappingTemplate:
    """UK / UKGBC COBie template — maps BS1192-4 and UKGBC psets."""
    base = _standard_template()
    extra = [
        PropertyMapping("Pset_Asset",           "AssetTag",         "Component", "TagNumber"),
        PropertyMapping("Pset_Asset",           "AssetIdentifier",  "Component", "Name"),
        PropertyMapping("Pset_Asset",           "ScheduleWork",     "Job", "Category"),
        PropertyMapping("Pset_Facility",        "FacilityReference","Facility", "Name"),
        PropertyMapping("Pset_Facility",        "GrossFloorArea",   "Facility", "GrossArea"),
        PropertyMapping("Pset_Facility",        "NetInternalArea",  "Facility", "SiteName"),
        PropertyMapping("Pset_SpaceOccupancy",  "OccupancyType",    "Space", "Category"),
        PropertyMapping("Pset_SpaceOccupancy",  "OccupancyNumber",  "Space", "Description"),
        PropertyMapping("Pset_Zone",            "ZoneCategory",     "Zone", "Category"),
    ]
    return MappingTemplate(
        template_name="uk_ukgbc",
        mappings=base.mappings + extra,
    )


def _singapore_corenet_template() -> MappingTemplate:
    """Singapore BCA CorNet COBie variant — e-Submission lifecycle extension."""
    base = _standard_template()
    extra = [
        PropertyMapping("Pset_BCASubmission",   "PermitNumber",    "Facility", "Description"),
        PropertyMapping("Pset_BCASubmission",   "PermitType",      "Facility", "Category"),
        PropertyMapping("Pset_BCABuilding",     "GrossFloorArea",  "Facility", "GrossArea"),
        PropertyMapping("Pset_BCABuilding",     "NetFloorArea",    "Facility", "SiteName"),
        PropertyMapping("Pset_BCABuilding",     "BuildingHeight",  "Floor", "Height"),
        PropertyMapping("Pset_BCASpace",        "RoomUse",         "Space", "Category"),
        PropertyMapping("Pset_BCASpace",        "RoomArea",        "Space", "GrossArea"),
        PropertyMapping("Pset_BCAAsset",        "AssetTag",        "Component", "TagNumber"),
        PropertyMapping("Pset_BCAAsset",        "MaintenanceCycle","Job", "Frequency"),
    ]
    return MappingTemplate(
        template_name="singapore_corenet",
        mappings=base.mappings + extra,
    )


_TEMPLATES: dict[str, MappingTemplate] = {}


def _init_templates() -> None:
    global _TEMPLATES
    if not _TEMPLATES:
        _TEMPLATES = {
            "standard":          _standard_template(),
            "federal_us":        _federal_us_template(),
            "uk_ukgbc":          _uk_ukgbc_template(),
            "singapore_corenet": _singapore_corenet_template(),
        }


def get_standard_template(name: str) -> MappingTemplate:
    """Return a named built-in mapping template.

    Parameters
    ----------
    name:
        One of ``"standard"``, ``"federal_us"``, ``"uk_ukgbc"``,
        ``"singapore_corenet"``.

    Raises
    ------
    ValueError
        If *name* is not recognised.
    """
    _init_templates()
    if name not in _TEMPLATES:
        raise ValueError(
            f"Unknown template {name!r}. "
            f"Available: {sorted(_TEMPLATES.keys())}"
        )
    return _TEMPLATES[name]


# ---------------------------------------------------------------------------
# IFC data extraction helpers
# ---------------------------------------------------------------------------

def _get_pset_value(
    ifc_data: dict[str, Any],
    pset_name: str,
    prop_name: str,
) -> str | None:
    """Extract a property value from a normalised IFC-data dict.

    The dict may be structured as:
    ``{"property_sets": {"<pset>": {"<prop>": "<value>", ...}, ...}, ...}``
    or a list of element dicts each with the same shape.
    """
    # Top-level pset dict
    psets = ifc_data.get("property_sets", {})
    if pset_name in psets:
        val = psets[pset_name].get(prop_name)
        if val is not None:
            return str(val)

    # Flatten from elements list
    elements = ifc_data.get("elements", [])
    for elem in elements:
        elem_psets = elem.get("property_sets", {})
        if pset_name in elem_psets:
            val = elem_psets[pset_name].get(prop_name)
            if val is not None:
                return str(val)

    return None


def _extract_elements_for_sheet(
    ifc_data: dict[str, Any],
    sheet_name: str,
    template: MappingTemplate,
) -> list[CobieRow]:
    """Build COBie rows for *sheet_name* from *ifc_data* using *template*."""
    # Gather all mappings that target this sheet
    sheet_mappings = [m for m in template.mappings if m.cobie_sheet == sheet_name]
    if not sheet_mappings:
        return []

    # Get all relevant IFC elements (or treat top-level as a single element)
    elements: list[dict[str, Any]] = ifc_data.get("elements", [{}])
    if not elements:
        elements = [ifc_data]

    # IFC-class → COBie sheet affinity
    _ifc_sheet_map = {
        "IfcBuilding": "Facility",
        "IfcBuildingStorey": "Floor",
        "IfcSpace": "Space",
        "IfcZone": "Zone",
        "IfcTypeObject": "Type",
        "IfcElement": "Component",
        "IfcSystem": "System",
        "IfcPerson": "Contact",
        "IfcOrganization": "Contact",
    }

    rows: list[CobieRow] = []
    now_iso = "2026-06-03T00:00:00"
    created_by = ifc_data.get("created_by", "n/a")

    # Build synthetic "top-level" element from root property_sets if no
    # elements cover the top-level psets (e.g. Contact comes from the
    # top-level Pset_ContactInformation).
    synthetic_root = {"property_sets": ifc_data.get("property_sets", {})}

    for elem in elements + [synthetic_root]:
        ifc_class = elem.get("ifc_class", "")
        # Only include elements that affinity-match this sheet,
        # or include all if no ifc_class (synthetic / top-level element)
        affiliated_sheet = _ifc_sheet_map.get(ifc_class)
        if affiliated_sheet and affiliated_sheet != sheet_name:
            continue
        if not ifc_class and sheet_name not in ("Contact", "Facility", "Job", "Document", "Zone", "Attribute"):
            # Synthetic root is only used for sheets without a direct IFC class
            continue

        row: CobieRow = {
            "Name":      elem.get("name") or elem.get("Name") or f"{sheet_name}-{uuid.uuid4().hex[:8]}",
            "CreatedBy": created_by,
            "CreatedOn": elem.get("created_on", now_iso),
            "Category":  elem.get("category", "n/a"),
        }

        # Apply mappings
        for mapping in sheet_mappings:
            # Try element-level psets first, then top-level
            val = None
            elem_psets = elem.get("property_sets", {})
            if mapping.ifc_pset_name in elem_psets:
                val = elem_psets[mapping.ifc_pset_name].get(mapping.ifc_property_name)
            if val is None:
                val = _get_pset_value(ifc_data, mapping.ifc_pset_name, mapping.ifc_property_name)
            if val is None and mapping.default_value:
                val = mapping.default_value
            if val is not None:
                row[mapping.cobie_column] = str(val)

        rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# apply_mapping_template
# ---------------------------------------------------------------------------

def apply_mapping_template(
    ifc_data: dict[str, Any],
    template: MappingTemplate,
) -> CobieDeliverable:
    """Map *ifc_data* through *template* and return a :class:`CobieDeliverable`.

    Parameters
    ----------
    ifc_data:
        Normalised IFC data dict.  Shape::

            {
              "created_by": "user@example.com",
              "property_sets": {"PsetName": {"PropName": "value", ...}, ...},
              "elements": [
                  {
                    "ifc_class": "IfcSpace",
                    "name": "Room 101",
                    "property_sets": {...},
                  },
                  ...
              ],
            }

    template:
        A :class:`MappingTemplate` (built-in or custom).

    Returns
    -------
    CobieDeliverable
        All 18 COBie sheets populated from the IFC data.
    """
    sheets: list[CobieSheet] = []
    for sheet_name in COBIE_SHEET_NAMES:
        req_cols = REQUIRED_COLUMNS.get(sheet_name, [])
        rows = _extract_elements_for_sheet(ifc_data, sheet_name, template)
        # Collect all columns present in rows + required columns
        col_set: list[str] = []
        seen: set[str] = set()
        for col in req_cols:
            if col not in seen:
                col_set.append(col)
                seen.add(col)
        for row in rows:
            for col in row:
                if col not in seen:
                    col_set.append(col)
                    seen.add(col)
        sheets.append(CobieSheet(name=sheet_name, columns=col_set, rows=rows))

    return CobieDeliverable(sheets=sheets)


# ---------------------------------------------------------------------------
# validate_cobie_deliverable
# ---------------------------------------------------------------------------

def validate_cobie_deliverable(deliverable: CobieDeliverable) -> list[str]:
    """Validate a :class:`CobieDeliverable` for COBie 2.4 compliance.

    Checks
    ------
    - All 18 COBie sheets present
    - Required columns present in each sheet's ``columns`` list
    - GUIDs (any column ending in ``ID`` or ``Guid``) unique within their sheet

    Returns
    -------
    list[str]
        List of human-readable error messages.  Empty → deliverable is valid.
    """
    errors: list[str] = []

    present_sheets = {s.name for s in deliverable.sheets}

    # Check all 18 sheets present
    for name in COBIE_SHEET_NAMES:
        if name not in present_sheets:
            errors.append(f"Missing required sheet: {name}")

    for sheet in deliverable.sheets:
        req_cols = REQUIRED_COLUMNS.get(sheet.name, [])
        sheet_col_set = set(sheet.columns)

        # Required column presence
        for col in req_cols:
            if col not in sheet_col_set:
                errors.append(
                    f"Sheet '{sheet.name}': missing required column '{col}'"
                )

        # GUID uniqueness
        guid_cols = [c for c in sheet.columns if c.lower().endswith(("id", "guid", "email"))]
        for gcol in guid_cols:
            seen_vals: set[str] = set()
            for row in sheet.rows:
                val = row.get(gcol, "").strip()
                if val and val not in ("n/a", ""):
                    if val in seen_vals:
                        errors.append(
                            f"Sheet '{sheet.name}': duplicate value '{val}' "
                            f"in column '{gcol}'"
                        )
                    seen_vals.add(val)

    return errors


# ---------------------------------------------------------------------------
# compute_completeness_score
# ---------------------------------------------------------------------------

def compute_completeness_score(deliverable: CobieDeliverable) -> float:
    """Compute COBie completeness as a fraction of required columns populated.

    For each sheet: count required columns that have at least one non-empty,
    non-"n/a" value in any row.  Returns 0.0–1.0.
    """
    total_required = 0
    total_populated = 0

    for sheet in deliverable.sheets:
        req_cols = REQUIRED_COLUMNS.get(sheet.name, [])
        total_required += len(req_cols)
        for col in req_cols:
            # A column counts as populated if at least one row has a value
            populated = any(
                row.get(col, "").strip() not in ("", "n/a")
                for row in sheet.rows
            )
            if populated:
                total_populated += 1

    if total_required == 0:
        return 1.0
    return total_populated / total_required


# ---------------------------------------------------------------------------
# export_cobie_excel
# ---------------------------------------------------------------------------

def export_cobie_excel(deliverable: CobieDeliverable, output_path: str) -> str:
    """Write *deliverable* to an .xlsx file at *output_path*.

    Uses ``openpyxl`` when available; falls back to a multi-sheet CSV bundle
    (one .csv per sheet, written to ``output_path + ".<SheetName>.csv"``) if
    openpyxl is absent.

    Returns
    -------
    str
        Absolute path of the written file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        _has_openpyxl = True
    except ImportError:
        _has_openpyxl = False

    if _has_openpyxl:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if wb.worksheets:
            wb.remove(wb.active)  # type: ignore[arg-type]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center")

        for sheet in deliverable.sheets:
            ws = wb.create_sheet(title=sheet.name[:31])  # Excel tab limit 31 chars
            # Write header row
            for col_idx, col_name in enumerate(sheet.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = max(12, len(col_name) + 2)
            # Write data rows
            for row_idx, row in enumerate(sheet.rows, start=2):
                for col_idx, col_name in enumerate(sheet.columns, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

        wb.save(output_path)
        return output_path

    else:
        # Fallback: one CSV per sheet
        import pathlib
        base = pathlib.Path(output_path)
        written: list[str] = []
        for sheet in deliverable.sheets:
            csv_path = str(base) + f".{sheet.name}.csv"
            with open(csv_path, "w", newline="", encoding="utf-8") as fh:
                writer = csv.DictWriter(fh, fieldnames=sheet.columns, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(sheet.rows)
            written.append(csv_path)
        return written[0] if written else output_path


# ---------------------------------------------------------------------------
# export_cobie_xml
# ---------------------------------------------------------------------------

def export_cobie_xml(deliverable: CobieDeliverable, output_path: str) -> str:
    """Write *deliverable* to a COBie XML file at *output_path*.

    The XML structure follows the buildingSMART COBie-XML schema::

        <COBie version="2.4">
          <Facility>
            <Contact>...</Contact>
            ...
          </Facility>
        </COBie>

    Returns
    -------
    str
        Absolute path of the written file.
    """
    root = ET.Element("COBie")
    root.set("version", "2.4")
    root.set("xmlns", "http://docs.buildingsmartalliance.org/nbims03/cobie/cobie.xsd")

    for sheet in deliverable.sheets:
        sheet_el = ET.SubElement(root, sheet.name)
        for row in sheet.rows:
            row_el = ET.SubElement(sheet_el, "Row")
            for col_name, col_val in row.items():
                col_el = ET.SubElement(row_el, col_name.replace(" ", "_"))
                col_el.text = col_val

    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")
    tree.write(output_path, encoding="utf-8", xml_declaration=True)
    return output_path
